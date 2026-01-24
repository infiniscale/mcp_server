"""Excel 引擎封装 - 使用 openpyxl 解析 Excel 文件。"""

import csv
import time
from pathlib import Path
from typing import Any, Dict, List


async def convert_with_excel(
    file_path: str,
    work_dir: Path = None,
) -> Dict[str, Any]:
    """
    将 Excel/CSV 文件转换为 Markdown 表格。

    Args:
        file_path: 输入文件路径
        work_dir: 工作目录

    Returns:
        Dict[str, Any]: 转换结果
    """
    start_time = time.time()
    attempt = {
        "engine": "excel",
        "status": "running",
        "error_code": None,
        "error_message": None,
        "elapsed_ms": 0,
    }

    warnings = []
    file_path = Path(file_path)
    ext = file_path.suffix.lower()

    try:
        if ext == ".csv":
            markdown_text = _convert_csv_to_markdown(file_path)
        elif ext == ".xlsx":
            markdown_text = _convert_xlsx_to_markdown(file_path)
        elif ext == ".xls":
            # xls 格式提示需要转换
            attempt["status"] = "error"
            attempt["error_code"] = "E_XLS_NOT_SUPPORTED"
            attempt["error_message"] = "xls 格式暂不直接支持，请先转换为 xlsx 格式"
            attempt["elapsed_ms"] = int((time.time() - start_time) * 1000)

            return {
                "ok": False,
                "attempt": attempt,
                "error_code": "E_XLS_NOT_SUPPORTED",
                "error_message": "xls 格式暂不直接支持，请先转换为 xlsx 格式（可使用 LibreOffice: soffice --headless --convert-to xlsx）",
                "warnings": ["建议安装 LibreOffice 并使用 soffice --headless --convert-to xlsx 转换"]
            }
        else:
            attempt["status"] = "error"
            attempt["error_code"] = "E_TYPE_UNSUPPORTED"
            attempt["error_message"] = f"Excel 引擎不支持的文件类型: {ext}"
            attempt["elapsed_ms"] = int((time.time() - start_time) * 1000)

            return {
                "ok": False,
                "attempt": attempt,
                "error_code": "E_TYPE_UNSUPPORTED",
                "error_message": f"Excel 引擎不支持的文件类型: {ext}",
                "warnings": warnings
            }

        attempt["status"] = "success"
        attempt["elapsed_ms"] = int((time.time() - start_time) * 1000)

        return {
            "ok": True,
            "markdown_text": markdown_text,
            "output_dir": str(work_dir / "output") if work_dir else None,
            "files": [],
            "warnings": warnings,
            "attempt": attempt
        }

    except ImportError as e:
        attempt["status"] = "error"
        attempt["error_code"] = "E_OPENPYXL_NOT_INSTALLED"
        attempt["error_message"] = f"openpyxl 未安装: {str(e)}"
        attempt["elapsed_ms"] = int((time.time() - start_time) * 1000)

        return {
            "ok": False,
            "attempt": attempt,
            "error_code": "E_OPENPYXL_NOT_INSTALLED",
            "error_message": "openpyxl 未安装，请运行: pip install openpyxl",
            "warnings": ["请安装 openpyxl: pip install openpyxl"]
        }

    except Exception as e:
        attempt["status"] = "error"
        attempt["error_code"] = "E_EXCEL_FAILED"
        attempt["error_message"] = str(e)
        attempt["elapsed_ms"] = int((time.time() - start_time) * 1000)

        return {
            "ok": False,
            "attempt": attempt,
            "error_code": "E_EXCEL_FAILED",
            "error_message": str(e),
            "warnings": warnings
        }


def _convert_csv_to_markdown(file_path: Path) -> str:
    """将 CSV 文件转换为 Markdown 表格。"""
    markdown_parts = []

    # 尝试不同的编码
    encodings = ["utf-8", "utf-8-sig", "gbk", "latin-1"]

    for encoding in encodings:
        try:
            with open(file_path, "r", encoding=encoding, newline="") as f:
                # 尝试检测分隔符
                sample = f.read(8192)
                f.seek(0)

                # 简单的分隔符检测
                if "\t" in sample and sample.count("\t") > sample.count(","):
                    delimiter = "\t"
                else:
                    delimiter = ","

                reader = csv.reader(f, delimiter=delimiter)
                rows = list(reader)

                if not rows:
                    return "*(空表格)*"

                # 转换为 Markdown 表格
                markdown_parts.append(f"# {file_path.stem}\n")
                markdown_parts.append(_rows_to_markdown_table(rows))

                return "\n".join(markdown_parts)

        except UnicodeDecodeError:
            continue

    raise ValueError(f"无法解析 CSV 文件，尝试的编码: {encodings}")


def _convert_xlsx_to_markdown(file_path: Path) -> str:
    """将 XLSX 文件转换为 Markdown。"""
    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxl 未安装")

    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    markdown_parts = []

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]

        # 获取所有行数据
        rows = []
        for row in sheet.iter_rows(values_only=True):
            # 转换为字符串，处理 None 值
            str_row = [str(cell) if cell is not None else "" for cell in row]
            # 跳过全空行
            if any(cell.strip() for cell in str_row):
                rows.append(str_row)

        if not rows:
            markdown_parts.append(f"## {sheet_name}\n\n*(空表格)*\n")
            continue

        # 添加 sheet 标题
        markdown_parts.append(f"## {sheet_name}\n")
        markdown_parts.append(_rows_to_markdown_table(rows))
        markdown_parts.append("")

    wb.close()

    if not markdown_parts:
        return "*(空工作簿)*"

    return "\n".join(markdown_parts)


def _rows_to_markdown_table(rows: List[List[str]]) -> str:
    """将行数据转换为 Markdown 表格。"""
    if not rows:
        return ""

    # 计算每列的最大宽度
    num_cols = max(len(row) for row in rows)

    # 标准化行（确保每行有相同的列数）
    normalized_rows = []
    for row in rows:
        normalized_row = list(row) + [""] * (num_cols - len(row))
        # 清理单元格内容（移除换行符，限制长度）
        normalized_row = [_clean_cell(cell) for cell in normalized_row]
        normalized_rows.append(normalized_row)

    # 生成表格
    lines = []

    # 表头（第一行）
    header = normalized_rows[0]
    lines.append("| " + " | ".join(header) + " |")

    # 分隔行
    lines.append("| " + " | ".join(["---"] * num_cols) + " |")

    # 数据行
    for row in normalized_rows[1:]:
        lines.append("| " + " | ".join(row) + " |")

    return "\n".join(lines)


def _clean_cell(cell: str) -> str:
    """清理单元格内容。"""
    if not cell:
        return ""

    # 移除换行符
    cell = cell.replace("\n", " ").replace("\r", "")

    # 转义管道符
    cell = cell.replace("|", "\\|")

    # 限制长度
    if len(cell) > 100:
        cell = cell[:97] + "..."

    return cell.strip()


def is_excel_available() -> bool:
    """检查 Excel 引擎是否可用。"""
    try:
        import openpyxl
        return True
    except ImportError:
        return False
